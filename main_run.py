from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox
from window_ui import Ui_MainWindow
from PySide6.QtGui import QImage, QPixmap
from PySide6.QtCore import QTimer

import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pathlib

from fer import FER

import pandas as pd
import cv2
import sys
import os


class Program_Window(QMainWindow, Ui_MainWindow):
    def __init__(self):
        QMainWindow.__init__(self)

        self.cam = cv2.VideoCapture(0)
        self.img_counter = 0
        self.camera_active = True
        self.capture_started = False
        self.remaining_time = 0

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.show()

        self.ui.stackedWidget.setCurrentWidget(self.ui.front_page)  # Set initial page to front_page

        # QTimer to switch to camera_page
        QTimer.singleShot(2000, self.switch_to_camera_page)
        
        self.createdatafile()

        self.ui.btn_capture.clicked.connect(self.capture_button_clicked)
        
        #Combo Box update the timer countdown
        self.ui.cb_timer.currentIndexChanged.connect(self.update_remaining_time)
        
        
        self.ui.btn_retake.clicked.connect(self.cancel_capture)
        self.ui.btn_saveIMG.clicked.connect(self.save_image)


        self.timer_label_update = QTimer()
        self.timer_label_update.timeout.connect(self.update_label_text)
        
        
    def perform_face_and_emotion_detection(self, img_path):
        frame = cv2.imread(img_path)
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        cascade_path = pathlib.Path(cv2.__file__).parent.absolute() / "data/haarcascade_frontalface_default.xml"
        clf = cv2.CascadeClassifier(str(cascade_path))

        faces = clf.detectMultiScale(
            gray,
            scaleFactor=1.1,
            minNeighbors=5,
            minSize=(30, 30),
            flags=cv2.CASCADE_SCALE_IMAGE
        )

        for (x, y, width, height) in faces:
            cv2.rectangle(frame, (x, y), (x + width, y + height), (255, 255, 0), 2)
            
            break



        currentdate = datetime.today().strftime('%Y-%m-%d')
        wb = load_workbook('data_' + currentdate + '.xlsx')
        ws = wb.active
        
        
        emotion_detector = FER(mtcnn=True)
        analysis = emotion_detector.detect_emotions(frame)
        dominant_emotion, emotion_score = emotion_detector.top_emotion(frame)
        
        for x in range(1,2):
            currentdate = datetime.today().strftime('%m/%d/%y')
            ws.cell(row=x, column=1).value = currentdate


        for emote in dominant_emotion:
            if dominant_emotion == 'disgust':
                disgust = ws.cell(row=1, column=2).value
                ws['B1'] = disgust + 1
            elif dominant_emotion == 'fear':
                fear = ws.cell(row=1, column=2).value
                ws['B1'] = fear + 1
            elif dominant_emotion == 'angry':
                angry = ws.cell(row=1, column=2).value
                ws['B1'] = angry + 1
            elif dominant_emotion == 'sad':
                sad = ws.cell(row=1, column=3).value
                ws['C1'] = sad + 1
            elif dominant_emotion == 'neutral':
                neutral = ws.cell(row=1,column=4).value
                ws['D1'] = neutral + 1
            elif dominant_emotion == 'happy':
                happy = ws.cell(row=1, column=5).value
                ws['E1'] = happy + 1
            elif dominant_emotion == 'surprise':
                surprise = ws.cell(row=1, column=5).value
                ws['F1'] = surprise + 1

                break
            
            break
        



        for i in range(1,2):
            value1 = ws.cell(row=i, column=2).value
            value2 = ws.cell(row=i, column=3).value
            value3 = ws.cell(row=i, column=4).value
            value4 = ws.cell(row=i, column=5).value
            value5 = ws.cell(row=i, column=6).value
            totalvalue = value1+value2+value3+value4+value5
            ws.cell(row=i, column=7).value = totalvalue

            break


        wb.save('data_' + datetime.today().strftime('%Y-%m-%d') + '.xlsx')

        txt = str(dominant_emotion)
        cv2.putText(frame, txt, (50, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)

        qimage = self.array_to_qimage(frame)
        pixmap = QPixmap.fromImage(qimage)
        self.ui.lbl_image_preview.setPixmap(pixmap)
        self.ui.lbl_image_preview.setScaledContents(True)

    def saveExcelAsCsvFile(self, excel_file_path, csv_file_path):
        try:
            # Load the Excel file using pandas
            df = pd.read_excel(excel_file_path)
            
            # Convert and save the DataFrame to CSV
            df.to_csv(csv_file_path, index=False)

            QMessageBox.information(self, "Info", "Excel file converted and saved as CSV successfully")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save Excel as CSV: {e}")


    def createdatafile(self):
        wb = openpyxl.Workbook()
        ws = wb.active


        ws ['B1'] = 0
        ws ['C1'] = 0
        ws ['D1'] = 0
        ws ['E1'] = 0
        ws ['F1'] = 0
        ws ['G1'] = 0



        for x in range(1,2):
            currentdate = datetime.today().strftime('%m/%d/%y')
            ws.cell(row=x, column=1).value = currentdate


        for i in range(1,2):
            value1 = ws.cell(row=i, column=2).value
            value2 = ws.cell(row=i, column=3).value
            value3 = ws.cell(row=i, column=4).value
            value4 = ws.cell(row=i, column=5).value
            value5 = ws.cell(row=i, column=6).value
            totalvalue = value1+value2+value3+value4+value5
            ws.cell(row=i, column=7).value = totalvalue

        wb.save('data_'+datetime.today().strftime('%Y-%m-%d')+'.xlsx')

    def show_camera(self):
        cascade_path = pathlib.Path(cv2.__file__).parent.absolute() / "data/haarcascade_frontalface_default.xml"
        clf = cv2.CascadeClassifier(str(cascade_path))
        while self.camera_active:  # Loop until camera_active is False
            ret, frame = self.cam.read()

            if not ret:
                print("Failed to grab frame")
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = clf.detectMultiScale(
                gray,
                scaleFactor=1.1,
                minNeighbors=5,
                minSize=(30, 30),
                flags=cv2.CASCADE_SCALE_IMAGE
            )

            for (x, y, width, height) in faces:
                cv2.rectangle(frame, (x, y), (x + width, y + height), (255, 255, 0), 2)

            self.display_frame(frame)

            k = cv2.waitKey(1)

            if k % 256 == 27:
                print("Escape Hit, closing the app")
                self.camera_active = False  # Set camera_active to False to exit the loop

        self.cam.release()
        cv2.destroyAllWindows()

    def display_frame(self, frame):
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        height, width, channel = frame.shape
        bytes_per_line = 3 * width
        q_image = QImage(frame.data, width, height, bytes_per_line, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(q_image)
        self.ui.imgLabel.setPixmap(pixmap)
        self.ui.imgLabel.setScaledContents(True)

    @staticmethod
    def array_to_qimage(arr):
        height, width, channel = arr.shape
        bytes_per_line = 3 * width
        qimage = QImage(arr.data, width, height, bytes_per_line, QImage.Format_RGB888)
        return qimage.rgbSwapped()

    def capture_image(self):
        ret, frame = self.cam.read()
        if ret:
            img_name = "Photo.jpg"
            cv2.imwrite(img_name, frame)
            self.img_counter += 1
            self.ui.stackedWidget.setCurrentWidget(self.ui.preview_page)
            self.display_captured_image(img_name)
            self.perform_face_and_emotion_detection(img_name)
            
            current_date = datetime.today().strftime('%Y-%m-%d')
            excel_file_path = 'data_' + current_date + '.xlsx'
            csv_file_path = 'data_' + current_date + '.csv'
            self.saveExcelAsCsvFile(excel_file_path, csv_file_path)
                    
    def display_captured_image(self, img_path):
        frame = cv2.imread(img_path)
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        cascade_path = pathlib.Path(cv2.__file__).parent.absolute() / "data/haarcascade_frontalface_default.xml"
        clf = cv2.CascadeClassifier(str(cascade_path))

        faces = clf.detectMultiScale(
            gray,
            scaleFactor=1.1,
            minNeighbors=5,
            minSize=(30, 30),
            flags=cv2.CASCADE_SCALE_IMAGE
        )

        for (x, y, width, height) in faces:
            cv2.rectangle(frame, (x, y), (x + width, y + height), (255, 255, 0), 2)

        qimage = self.array_to_qimage(frame)
        pixmap = QPixmap.fromImage(qimage)
        self.ui.lbl_image_preview.setPixmap(pixmap)
        self.ui.lbl_image_preview.setScaledContents(True)

    def switch_to_camera_page(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.camera_page)
        self.show_camera()

    def capture_button_clicked(self):
        delay_milliseconds = self.remaining_time * 1000  # Convert seconds to milliseconds
        QTimer.singleShot(delay_milliseconds, self.capture_image)

        # Start the QTimer to update the label text periodically
        self.timer_label_update.start(1000)  # Update every second (1000 milliseconds)
        
        # Perform face and emotion detection immediately after capture
        self.perform_face_and_emotion_detection("Photo.jpg")
        
    # Updating the text base on the time left
    def update_label_text(self):
        self.remaining_time -= 1
        self.ui.lbl_time_count.setText(str(self.remaining_time))
        if self.remaining_time <= 0:
            self.timer_label_update.stop()

    def update_remaining_time(self):
        selected_timer = self.ui.cb_timer.currentText()
        self.remaining_time = int(selected_timer)
        self.ui.lbl_time_count.setText(str(self.remaining_time))

    def cancel_capture(self):
        
        # Go back to the camera page
        self.ui.stackedWidget.setCurrentWidget(self.ui.camera_page)
        
        # Delete the last captured image if it exists
        if self.img_counter > 0:
            img_name = "Photo.jpg"
            os.remove(img_name)
            self.img_counter -= 1
            self.remaining_time = 3
            self.ui.lbl_time_count.setText(str(self.remaining_time))
            self.ui.cb_timer.setCurrentIndex(2)

    def save_image(self):
        if self.img_counter > 0:
            img_name = "Photo.jpg"
            
            # Perform the save operation here
            # You can modify the image to the desired location
            
            destination_path = "C:/"
            
            frame = cv2.imread(img_name)
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            cascade_path = pathlib.Path(cv2.__file__).parent.absolute() / "data/haarcascade_frontalface_default.xml"
            clf = cv2.CascadeClassifier(str(cascade_path))

            faces = clf.detectMultiScale(
                gray,
                scaleFactor=1.1,
                minNeighbors=5,
                minSize=(30, 30),
                flags=cv2.CASCADE_SCALE_IMAGE
            )

            for (x, y, width, height) in faces:
                cv2.rectangle(frame, (x, y), (x + width, y + height), (255, 255, 0), 2)

            qimage = self.array_to_qimage(frame)
            pixmap = QPixmap.fromImage(qimage)
            self.ui.lbl_image_preview.setPixmap(pixmap)
            self.ui.lbl_image_preview.setScaledContents(True)

            cv2.putText(frame, self.ui.lbl_image_preview.text(), (50, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)
            
            emotion_detector = FER(mtcnn=True)
            analysis = emotion_detector.detect_emotions(frame)
            dominant_emotion, emotion_score = emotion_detector.top_emotion(frame)

            txt = str(dominant_emotion)
            cv2.putText(frame, txt, (50, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)

            cv2.imwrite(destination_path + img_name, frame)
            
            QMessageBox.information(self, "Info", "Photo Saved Successfully") 
            
            self.ui.stackedWidget.setCurrentWidget(self.ui.camera_page)
            self.remaining_time = 3
            self.ui.lbl_time_count.setText(str(self.remaining_time))
            self.ui.cb_timer.setCurrentIndex(2)
            
    def closeEvent(self, event):
        self.camera_active = False
        self.cam.release()
        event.accept()

if __name__ == "__main__":
    self = Program_Window
    self.createdatafile(self)
    app = QApplication(sys.argv)
    window = Program_Window()
    window.setWindowTitle("Camera")
    window.show()
    sys.exit(app.exec())